import openpyxl
import string

#후보자 정보
class Candidate:
    def __init__(self, _region, _party, _name):
        self.region = _region
        self.party = _party      #후보자 정당명
        self.name = _name        #후보자 이름
        
        self.in_prevote = 0      #관내 사전투표 득표수
        self.out_prevote = 0     #관외 사전투표 득표수
        
        self.in_prevoteRate = 0  #관내 사전투표 비율
        self.out_prevoteRate = 0 #관외 사전투표 비율

        self.matchRate = 0       #관내사전, 관외사전 일치율

    def InputInPrevote(self, n):
        self.in_prevote += n

    def InputOutPrevote(self, n):
        self.out_prevote = n

    def InputInPrevoteRate(self, n):
        self.in_prevoteRate = n
        
    def InputOutPrevoteRate(self, n):
        self.out_prevoteRate = n

    def InputMatchRate(self, n):
        self.matchRate = n

    def AddResult(self):
        list = [self.region, self.party, self.name, self.in_prevote,
                self.out_prevote, self.in_prevoteRate, self.out_prevoteRate,
                self.matchRate]
        result.append(list)


#관내 사전투표 비율
#parameter: A, B 후보자
#return: A 후보자의 관내사전투표수 / A, B 후보자의 관내사전특표수
#        B 후보자의 관내사전특표수 / A, B 후보자의 관내사전득표수
def InPrevoteRate(a, b):
    total = a.in_prevote + b.in_prevote
    return (a.in_prevote / total), (b.in_prevote / total)

#관외 사전투표 비율
#parameter: A, B 후보자
#return: A 후보자의 관외사전득표수 / A, B 후보자의 관외사전득표수
#        B 후보자의 관외사전득표수 / A, B 후보자의 관외사전득표
def OutPrevoteRate(a, b):
    total = a.out_prevote + b.out_prevote
    return (a.out_prevote / total), (b.out_prevote / total)


#관내, 관내 사전투표 일치율
#parameter: A 후보자
#return: (사전득표수 / 사전득표수) %
def CalculateMatchRate(a):
    return min(a.in_prevoteRate, a.out_prevoteRate) / max(a.in_prevoteRate, a.out_prevoteRate)


candidateList = []
candidate = [0,0]
result = [["지역구", "정당", "후보자명", "관내 사전득표수", "관외 사전득표수",
          "관내 사전득표율", "관외 사전득표율", "일치율"]]


#불러올 엑셀 파일 이름
file_region_list = ["강원", "경기", "경남", "경북", "광주", "대구",
                    "대전", "부산", "서울", "세종", "울산", "인천",
                    "전남", "전북", "제주", "충남", "충북"]

for r in range(len(file_region_list)):
    file_name = "20대\\20대총선지역구(" + file_region_list[r] + ").xlsx"
    file = openpyxl.load_workbook(file_name)
    file_sheet_names = file.sheetnames


    for sheet_name in file_sheet_names:
        sheet = file[sheet_name]
        cnt = 0

        if sheet_name == "통영고성":
            continue

        for i in sheet.rows:
            for j in range(len(candidate)):
                column = string.ascii_uppercase[4 + j]

                #지역마다 후보자 정보 입력         
                if i[0].value != None and i[0].value != "선거구명":
                    cnt += 1

                    if cnt > 2 and j != 1:
                        candidateList.append([candidate[0], candidate[1]])
                    
                    party_cellname = column + str(i[0].row - 2)
                    name_cellname = column + str(i[0].row - 1)

                    region = i[0].value
                    party = sheet[party_cellname].value
                    name = sheet[name_cellname].value
    
                    candidate[j] = Candidate(region, party, name)

                #관내 사전투표 총합
                if i[1].value == "관내사전투표":
                    in_cellname = column + str(i[1].row)
                    in_prevote = sheet[in_cellname].value

                    candidate[j].InputInPrevote(in_prevote)
            
                #관외 사전투표 총합
                if i[1].value == "관외사전투표":
                    out_cellname = column + str(i[1].row)
                    out_prevote = sheet[out_cellname].value

                    candidate[j].InputOutPrevote(out_prevote)


        
candidateList.append(candidate)

for i in range(1, len(candidateList)):
    #관내, 관외 사전득표율 계산
    a_inPrevoteRate, b_inPrevoteRate = InPrevoteRate(candidateList[i][0], candidateList[i][1])
    a_outPrevoteRate, b_outPrevoteRate = OutPrevoteRate(candidateList[i][0], candidateList[i][1])

    candidateList[i][0].InputInPrevoteRate(a_inPrevoteRate)
    candidateList[i][0].InputOutPrevoteRate(a_outPrevoteRate)

    candidateList[i][1].InputInPrevoteRate(b_inPrevoteRate)
    candidateList[i][1].InputOutPrevoteRate(b_outPrevoteRate)

    #각 후보자의 관내, 관외 사전득표율 일치율 계산
    for j in range(len(candidateList[i])):
        matchRate = CalculateMatchRate(candidateList[i][j])
        candidateList[i][j].InputMatchRate(matchRate)
        
        candidateList[i][j].AddResult()

            

#결과
result_file = openpyxl.Workbook()
result_sheet = result_file.active

for i in result:
    result_sheet.append(i)

result_file.save("20대 결과.xlsx")
